import openpyxl
from os import path
from PyQt5 import QtWidgets

class Planilha:

    def criar_planilha(self):
        if not path.exists('temp/info/material.xlsx'):
            planilha = openpyxl.Workbook()
            aba_ativa = planilha.create_sheet('material', 0)

            data = [
                ['Padrão', '7.1', 'Auto', '1/125', '0', '0', 'Auto', 'off'],
            ]

            # add column headings. NB. these must be strings
            aba_ativa.append(["Nome do Material", "Abertura", "ISO", "Vel. Disparo",
                              "WB Ajuste A", "WB Ajuste B", "Picture Style", "Comp. Exposição"])
            for row in data:
                aba_ativa.append(row)

            tabela = openpyxl.worksheet.table.Table(displayName="material", ref="A1:H2")

            # Add a default style with striped rows and banded columns
            estilo_tabela = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
            tabela.tableStyleInfo = estilo_tabela

            # Cria aba e salva tabela
            aba_ativa.add_table(tabela)
            planilha.save("temp/info/material.xlsx")

    def set_valor_planilha(self, valor1, valor2, valor3, valor4, valor5, valor6, valor7, valor8, local_planilha,
                           comboBox=None, tableWidget_planilha=None):
        planilha = openpyxl.load_workbook(local_planilha)
        aba_ativa = planilha.worksheets[0]

        linha = [valor1, valor2, valor3, valor4, valor5, valor6, valor7, valor8]

        aba_ativa.append(linha)
        ultima_linha = aba_ativa.max_row

        tabela = aba_ativa.tables['material']
        tabela.ref = f"A1:H{ultima_linha}"

        planilha.save(local_planilha)
        self.atualizar_planilha(tableWidget_planilha, local_planilha)
        self.atualizar_comboBox(comboBox, local_planilha)

    def excluir_linha(self, linha, local_planilha, comboBox=None, tableWidget_planilha=None):
        planilha = openpyxl.load_workbook(local_planilha)
        aba_ativa = planilha.worksheets[0]

        aba_ativa.delete_rows(linha)

        planilha.save(local_planilha)
        self.atualizar_planilha(tableWidget_planilha, local_planilha)
        self.atualizar_comboBox(comboBox, local_planilha)

    def atualizar_comboBox(self, comboBox, local_planilha):
        planilha = openpyxl.load_workbook(local_planilha)
        aba_ativa = planilha.worksheets[0]
        comboBox.clear()
        comboBox.addItems([''])

        for celula in aba_ativa['A']:
            if celula != aba_ativa['A1']:
                comboBox.addItems([celula.value])

    def atualizar_planilha(self, tableWidget_planilha, local_planilha):
        planilha = openpyxl.load_workbook(local_planilha)
        aba_ativa = planilha.worksheets[0]
        total_linhas = aba_ativa.max_row
        tableWidget_planilha.setRowCount(total_linhas)
        tableWidget_planilha.setColumnCount(8)
        for l in range(0, total_linhas):
            for c in range(0, 8):
                tableWidget_planilha.setItem(l, c, QtWidgets.QTableWidgetItem(aba_ativa.cell(l + 2, c + 1).value))