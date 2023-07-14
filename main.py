import ctypes
import os
import sys
import time
import openpyxl
from threading import Thread

from camera import Camera
from planilha import Planilha
from pastas import Pasta
from criptografia import Criptografar
from teste_rede import TesteRede
from imagem import Imagem
from PyQt5.QtWidgets import QMainWindow, QApplication
from PyQt5.QtGui import QPixmap, QIcon
from templates.designer import *
from templates.menu import *
from templates.material import *
from subprocess import call



#   Abre a parte de Designer Tela Principal
class App(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.menu_conf = None
        super().setupUi(self)
        ctypes.windll.kernel32.FreeConsole()

        Pasta.cria_pastas(self)
        self.criptografia = Criptografar()

        # Executa a função de ping no servidor de forma paralela
        Thread(target=self.teste_rede, daemon=True).start()

        # Testa câmera
        Thread(target=self.teste_camera, daemon=True).start()

        # Chama a classe planilha
        self.planilha = Planilha()
        self.planilha.criar_planilha()

        # Chama a funcão do botão capturar
        self.btnCapturar.clicked.connect(self.capturar)
        self.lineEdit_codigoDeBarras.returnPressed.connect(self.capturar)

        # Chama a funcão do botão voltar
        self.btnVoltar.clicked.connect(self.voltar)

        # Chama a funcão do botão salvar
        self.btnSalvar.clicked.connect(self.salvar_imagem)

        # Chama janela de menu
        self.actionConfigurar_Servidor.triggered.connect(self.menu_servidor)
        self.actionConfigurar_Material.triggered.connect(self.menu_material)

        # Chama função CROP
        self.btnCrop.clicked.connect(self.cortar_imagem)

        # Adicionar dados na comboBox
        self.comboBox_espessura.addItems(['1,5cm', '2cm', '3cm', '4cm', '5cm'])
        self.planilha.atualizar_comboBox(self.comboBox_material, r'temp/info/material.xlsx')
        self.comboBox_material.currentIndexChanged.connect(self.set_config_camera)

    def menu_servidor(self):
        self.menu_conf = AbreMenu_Servidor()
        self.menu_conf.show()

    def menu_material(self):
        self.menu_conf = AbreMenu_Material()
        self.menu_conf.showMaximized()

    def teste_rede(self):
        x = 0
        while True:
            teste = TesteRede.testeRede(self)
            if teste == 0 and x == 0:
                Imagem.backup_imagens_rede(self)
                x = 1
            if teste == 1:
                x = 0
            time.sleep(60)

    def teste_camera(self):
        while True:
            teste_camera = str(camera.detectar_camera())
            nome_camera = camera.nome_camera()
            if teste_camera == 'Conectada':
                self.btnCapturar.setIcon(QIcon('templates/icons/cameraON.png'))
                self.label_status.setText(f'Câmera {teste_camera}: {nome_camera}')
                self.label_status.setStyleSheet('QLabel {background-color: rgba(131, 131, 131, 50); color: #000000}')
                time.sleep(15)
            else:
                self.btnCapturar.setIcon(QIcon('templates/icons/cameraOFF.png'))
                self.label_status.setText('Câmera Desconectada')
                self.label_status.setStyleSheet('QLabel {background-color: rgba(131, 131, 131, 50); color: #000000}')
                time.sleep(3)

    def set_config_camera(self):
        if camera.detectar_camera() == 'Conectada':
            planilha_material = openpyxl.load_workbook('temp/info/material.xlsx')
            material = planilha_material['material']
            for linha in range(2, material.max_row + 1):
                if self.comboBox_material.currentText() == material.cell(linha, 1).value:
                    valor_abertura = material.cell(linha, 2).value
                    valor_iso = material.cell(linha, 3).value
                    valor_velDisparo = material.cell(linha, 4).value
                    valor_wba = material.cell(linha, 5).value
                    valor_wbb = material.cell(linha, 6).value
                    valor_pictureStyle = material.cell(linha, 7).value
                    valor_compExposicao = material.cell(linha, 8).value
                    camera.configurar_camera('aperture', valor_abertura)
                    camera.configurar_camera('iso', valor_iso)
                    camera.configurar_camera('shutterspeed', valor_velDisparo)
                    camera.configurar_camera('whitebalanceadjusta', valor_wba)
                    camera.configurar_camera('whitebalanceadjustb', valor_wbb)
                    camera.configurar_camera('picturestyle', valor_pictureStyle)
                    camera.configurar_camera('aeb', valor_compExposicao)
                    break

    # Função do botão capturar
    def capturar(self):
        if not os.path.exists('temp/img'):
            os.mkdir('temp/img')

        self.valor_required_comboBox(self.comboBox_material)
        self.valor_required_lineEdit(self.lineEdit_codigoDeBarras)
        if self.comboBox_material.currentText() and self.lineEdit_codigoDeBarras.text() != '':
            if camera.detectar_camera() == 'Conectada':

                camera.capturar_imagem()

                # Mosta a imagem capturada na tela
                self.imagem = r'temp/img/capt0000.jpg'
                self.imagem = QPixmap(self.imagem)
                self.labelImagem.setPixmap(self.imagem)
                self.cortar_imagem()

    def cortar_imagem(self):
        Imagem.cortar_imagem(self)

    def salvar_imagem(self):
        Imagem.salvar_imagem(self)


    def voltar(self):
        self.lineEdit_codigoDeBarras.clear()
        self.lineEdit_3.clear()
        self.lineEdit_7.clear()
        self.lineEdit_6.clear()
        self.lineEdit_8.clear()
        self.lineEdit_9.clear()

    def valor_required_lineEdit(self, campo):
        if campo.text() == '':
            campo.setStyleSheet('QLineEdit {background-color: rgb(255, 255, 255); '
                                'border-radius: 5%; padding: 5px; border: 1px solid red;}')
        else:
            campo.setStyleSheet('QLineEdit {border: none; border-radius: 5%; padding: 5px;'
                                            'border: 0.5px solid rgba(131, 131, 131, 50);'
                                            'border-bottom: 1px solid #838383;'
                                            'background-color: rgb(255, 255, 255);}'
                                
                                            'QLineEdit:focus {border: none;'
                                            'border: 0.5px solid rgba(131, 131, 131, 50);'
                                            'border-bottom: 1.5px solid #003ae5;}'
                                
                                            'QLineEdit:focus:hover {border: none;'
                                            'border: 0.5px solid rgba(131, 131, 131, 50);'
                                            'border-bottom: 1px solid #838383;}')

    def valor_required_comboBox(self, campo):
        if campo.currentText() == '':
            campo.setStyleSheet('QComboBox {background-color: rgb(255, 255, 255); border: 1px solid red;}')
        else:
            campo.setStyleSheet('QComboBox {background-color: rgb(255, 255, 255);}')


class AbreMenu_Servidor(QMainWindow, Ui_confServidor):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        self.criptografia = Criptografar()
        self.btnSalvar.clicked.connect(self.conectar_servidor)
        if os.path.exists(r'temp/info/mapping.txt'):
            with open(r'temp/info/mapping.txt', 'r') as file:
                file.seek(0)
                ler_ip = file.readline().strip()
                ler_usuario = file.readline().strip()
                self.inputServidor.setText(ler_ip)
                self.inputUsuario.setText(ler_usuario)
        if os.path.exists(r'temp/info/password.bin'):
            with open(r'temp/info/password.bin', 'rb') as file:
                file.seek(0)
                ler_senha = file.read()
                dec_ler_senha = self.criptografia.decodificar_valores(ler_senha)
                self.inputSenha.setText(dec_ler_senha)

    def conectar_servidor(self):
        ip_servidor = self.inputServidor.text()
        usuario = self.inputUsuario.text()
        senha = self.inputSenha.text()
        caminho_completo_servidor = fr'net use \\{ip_servidor} /user:{usuario} {senha}'
        conectado = call(caminho_completo_servidor)

        if conectado == 0 or ip_servidor == '':
            enc_senha = self.criptografia.codificar_valores(senha)
            dados = f'{ip_servidor}\n{usuario}'

            with open(r'temp/info/mapping.txt', 'w+') as file:
                file.write(dados)
            with open(r'temp/info/password.bin', 'wb') as file:
                file.write(enc_senha)
            self.labelStatusServidor.setText(f'Conectado com sucesso: {ip_servidor}')
            self.labelStatusServidor.setStyleSheet('QLabel {color: #228B22}')

        elif conectado == 1:
            self.labelStatusServidor.setText('Servidor não conectado: Informe os dados para conexão.')
            self.labelStatusServidor.setStyleSheet('QLabel {color: #FF0000}')
        else:
            self.labelStatusServidor.setText('Servidor não conectado: Os dados estão incorretos.')
            self.labelStatusServidor.setStyleSheet('QLabel {color: #FF0000}')


class AbreMenu_Material(QMainWindow, Ui_confMaterial):
    def __init__(self, parent=None):
        super().__init__(parent)
        super().setupUi(self)

        self.planilha = Planilha()
        self.planilha.criar_planilha()

        self.pushButton_add_linha.clicked.connect(self.adicionar_valor_planilha)
        self.pushButton_del_linha.clicked.connect(self.remover_valor_planilha)
        self.planilha.atualizar_planilha(self.tableWidget_material, r'temp/info/material.xlsx')
        self.lineEdit_nomeMaterial.returnPressed.connect(self.adicionar_valor_planilha)

        self.tableWidget_material.setColumnWidth(0,400)

    def adicionar_valor_planilha(self):
        valor1 = self.lineEdit_nomeMaterial.text()
        app.valor_required_lineEdit(self.lineEdit_nomeMaterial)
        valor2 = self.comboBox_Abertura.currentText()
        app.valor_required_comboBox(self.comboBox_Abertura)
        valor3 = self.comboBox_ISO.currentText()
        valor4 = self.comboBox_VelDisparo.currentText()
        app.valor_required_comboBox(self.comboBox_VelDisparo)
        valor5 = self.comboBox_WBa.currentText()
        valor6 = self.comboBox_WBb.currentText()
        valor7 = self.comboBox_PictureStyle.currentText()
        valor8 = self.comboBox_CompExposicao.currentText()
        local_planilha = r'temp/info/material.xlsx'
        comboBox = app.comboBox_material
        tableWidget_planilha = self.tableWidget_material

        if valor1 and valor2 and valor4 != '':
            self.planilha.set_valor_planilha(valor1, valor2, valor3, valor4, valor5, valor6, valor7,
                                         valor8, local_planilha, comboBox, tableWidget_planilha)

    def remover_valor_planilha(self):
        linha = self.tableWidget_material.currentRow()+2
        local_planilha = r'temp/info/material.xlsx'
        comboBox = app.comboBox_material
        tableWidget_planilha = self.tableWidget_material
        self.planilha.excluir_linha(linha,local_planilha, comboBox, tableWidget_planilha)


if __name__ == '__main__':
    camera = Camera()
    qt = QApplication(sys.argv)
    app = App()
    app.showMaximized()
    qt.exec_()
